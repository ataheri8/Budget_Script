#!/usr/bin/python

"""
This script has been created by Afsheen Taheri. This script is not
to be reproduced and shared without written consent.

The purpose of the script is to generate a budget based off user data.
The user is to enter their expenses and income to generate an organized
excel file that will help keep track of expenses and track how much is left
in the budget for the rest of the month.

Currently live:
- Able to create a new budget from scratch and add any purchases on to the newly
  created budget
- Able to add purchases to other budget excel files (in the same format as the
  one generated) instead of only being able to add it to freshly created budgets
- Color coding the "left in budget" costs to reflect how close the user is to reaching
  the limit to that category (< 90% - white meaning good, > 90% - yellow meaning close
  to limit, > 100% - red meaning limit has been exceeded

In development:
- A list of how many categories are in the green, yellow and red
"""


__author__ = "Afsheen Taheri"
__version__ = "1.0"
__email__ = "afsheen.t95@gmail.com"
__status__ = "Development"

import xlsxwriter
from openpyxl import load_workbook
import os
import sys
import time

"""
Defining the main menu of the program. Will call different functions depending
on what the user inputs
"""


def main_menu():
    os.system('cls')
    print("******Welcome to the budget creater******\n\n\n\n")
    print("1) Create a new budget")
    print("2) Add entries to existing budget")
    print("3) Exit program\n\n")
    choice = input("Enter your choice here: ")

    while True:
        if choice == '1':
            add_budget()
            break
        elif choice == '2':
            for file in os.listdir('.'):
                if file.endswith('.xlsx'):
                    print(file)
            choice = input("Which file would you like to edit? (no need to add the .xlsx at the end): ")
            edit_budget(file=(choice + ".xlsx"))
            break
        elif choice == '3':
            exit_program()
            break
        else:
            print("invalid choice, please try again")
            choice = input("Enter your choice here: ")

    
"""
Defining the add_budget function. This function's purpose is to take all the
user's data, such as their variable and non-variable expenses, the budget name,
monthly income and if any data needs to be added under the purchases. 
"""


def add_budget():
    os.system('cls')
    var_total = 0
    non_var_total = 0
    non_var = {}
    var = {}
    
    """
    Name the budget
    """
    budget_name = './' + input("Give your budget a name!: ") + '.xlsx'
    
    """
    Input for non variable expenses
    """
    print("\nPlease enter your non-variable expenses (such as rent, hydro, etc.(Format: 'Category/Cost') "
          "Type 'done' when finished or 'exit' to go back to the main menu\n")
    while True:
        entry = input()
        expense = entry.split("/")
        if entry == "done":
            break
        elif entry == "exit":
            non_var.clear()
            main_menu()
        elif len(expense) != 2:
            print("incorrect format. Please try again")
        else:
            try:
                non_var[expense[0]] = int(expense[1])
                non_var_total += int(expense[1])
            except ValueError:
                print("Please ensure you only use digits for the cost.")
            else:
                print("\nSuccessfully added! Next Entry: ")
            
    """
    Input for variable expenses
    """
    os.system('cls')
    print("\nPlease enter your variable expenses (such as entertainment, food, etc.(Format: 'Category/Cost') "
          "Type 'done' when finished or 'exit' to go back to the main menu\n")
    while True:
        entry = input()
        expense = entry.split("/")
        if entry == "done":
            break
        elif entry == "exit":
            var.clear()
            main_menu()
        elif len(expense) != 2:
            print("incorrect format. Please try again")
        else:
            try:
                var[expense[0]] = int(expense[1])
                var_total += int(expense[1])
            except ValueError:
                print("Please ensure you only use digits for the cost.")
            else:
                print("\nSuccessfully added! Next Entry:")
    """
    Input for monthly income
    """
    while True:
        try:
            income = int(input("Please input your monthly income: "))
            surplus = income - (var_total + non_var_total)
        except ValueError:
            print("only use numbers, no letters or special characters")
        else:
            if surplus < 0:
                user_input = input("This budget will carry a negative balance of ${}. "
                                   "Would you like to redo the budget?(yes/no): ".format(surplus))
                if user_input == "yes":
                    add_budget()
            break

    """
    Asks user if he would like to add any entries. If so, call the edit_budget
    function and then create the file. If no, then just create the file
    """
    print("Would you like to add any entries to this budget? (yes/no)")
    while True:
        choice = input()
        if choice == 'yes':
            try:
                create_file(budget_name, non_var, non_var_total, var, var_total, income, edit_budget(var))
            except:
                print("something went wrong, please try again")
                time.sleep(2)
                exit_program()
            else:
                print("Your budget has been created! Taking you back to the main menu...")
                time.sleep(2)
                main_menu()
        elif choice == 'no':
            try:
                create_file(budget_name, non_var, non_var_total, var, var_total, income)
            except:
                print("something went wrong, please try again")
                time.sleep(2)
                exit_program()
            else:
                print("Your budget has been created! Taking you back to the main menu...")
                time.sleep(2)
                main_menu()
        else:
            print("invalid choice, please enter a 'yes' or 'no'")


"""
Defining the edit_budget function. This function's purpose is to add purchase
entries to the budget. Currently can only add entries to budgets created
by the add_budget function.
"""

def edit_budget(var={}, file=""):
    os.system('cls')
    data = []
    temp_category = []
    categories = []
    row = 0

    if var != {}:
        temp_category = var

    if file != "":
        try:
            wb = load_workbook(file)
            ws = wb.active
        except FileNotFoundError:
            print("file was not found. Taking you back to the main menu...")
            time.sleep(2)
            main_menu()
        else:
            print("file has been successfully loaded!")
            time.sleep(1)

        for i in range(2, 50, 1):
            if ws['E{}'.format(i)].value is None:
                row = i
                break

        for i in range(2, 50, 1):
            if ws['C{}'.format(i)].value is None:
                break
            else:
                temp_category.append(ws['C{}'.format(i)].value)
    
    """
    Presents the variable categories that is in the user's budget
    """
    print("These are the current categories you have in your budget\n")
    for i in temp_category:
        print(i, end='  ')
        categories.append(i.capitalize())

    print("\n\nPlease add an entry (Format: 'Entry/Category/Cost/Date'). "
           "Ensure you pick one of your existing categories and the date "
           "format to be dd-mm-yyyy. Enter 'done' when finished")

    """
    Takes the user's input for their purchases
    """
    while True:
        entry = input()
        item = entry.split("/")
        if entry == 'done':
            break
        elif len(item) != 4:
            print("incorrect format. Please try again.")
        elif item[1].capitalize() not in categories:
            print("Please ensure you pick from one of the existing categories")
        else:
            try:
                if file != "":
                    ws['E{}'.format(row)] = item[0]
                    ws['F{}'.format(row)] = item[1]
                    ws['G{}'.format(row)] = float(item[2])
                    ws['H{}'.format(row)] = item[3]
                    row += 1
                else:
                    data.append([item[0], item[1], int(item[2]), item[3]])
            except ValueError:
                print("Please ensure you only use digits for the cost.")
            else:
                print("\nSuccessfully added! Next Entry:")

    if file != "":
        print("file has been successfully edited! Taking you back to the main menu...")
        wb.save(file)
        time.sleep(2)
        main_menu()
    else:
        return data
        
    
"""
Defining the exit_program function. This function's purpose is to display
a goodbye greeting and exit the program
"""


def exit_program():
    print("Have a great day!")
    time.sleep(2)
    sys.exit()


"""
Defining the create_file function. This function's purpose is to take the user data
passed by the add_budget and export the data into the excel file.
"""


def create_file(budget_name, non_var, non_var_total, var, var_total, income, data=[]):

    """
    Creating the excel file and takes the name passed by the user
    """
    total = 0
    extra_income = income - (var_total + non_var_total)
    workbook = xlsxwriter.Workbook(budget_name)
    worksheet_budget = workbook.add_worksheet('Budget')

    """
    Sets the formats to be used in the file
    """
    header = workbook.add_format(
        {
            "bg_color": "D8FFD0",
            "border": 0,
            "font": "Calibri",
            "font_size": 11,
            "align": "center",
            "valign": "vcenter",
        }
    )

    var_seperator = workbook.add_format(
        {
            "bg_color": "C5DCFF",
            "border": 0,
            "font": "Calibri",
            "font_size": 11,
            "align": "center",
            "valign": "vcenter",
        }
    )

    date = workbook.add_format(
        {
            'num_format': 'd-m-yyyy'
        }
    )

    num_field = workbook.add_format(
        {
            "num_format": 44
        }
    )

    caution = workbook.add_format(
        {
            "bg_color": "FDFFB4",
            "num_format": 44
        }
    )

    over = workbook.add_format(
        {
            "bg_color": "FFB5B4",
            "num_format": 44
        }
    )
    
    normal = workbook.add_format(
        {
            "font": "Calibri",
            "font_size": 11,
            "align": "center",
            "valign": "vcenter",
        }
    )

    worksheet_budget.conditional_format('D2:D100', {'type': 'blanks',
                                                    'stop_if_true': True,
                                                    'format': normal})

    """
        Sets the column sizes
        """
    worksheet_budget.set_column(
        "A:C",
        30
    )

    worksheet_budget.set_column(
        "D:D",
        15
    )

    worksheet_budget.set_column(
        "E:H",
        25
    )

    """
    Setting default values for rows, columns and counter
    """

    row = 0
    col = 0
    
    worksheet_budget.write_row(
        row, col, ["Expenses", "Cost of expenses", "Spending to date", "Date"],
        header
    )
    worksheet_budget.merge_range('C1:D1', 'Spending to date', header)
    worksheet_budget.merge_range('E1:H1', "Individual bought items", header)
    row += 1
    
    worksheet_budget.write_row(
        row, col, ["Non Variable"],
        var_seperator
    )

    for key, value in sorted(non_var.items()):
        row += 1
        worksheet_budget.write_row(row, col, [key.capitalize()], normal)
        worksheet_budget.write_row(row, col + 1, [value], num_field)

    row += 2

    worksheet_budget.write_row(row, col, ["Total Non Variable"], normal)
    worksheet_budget.write_row(row, col + 1, [non_var_total], num_field)

    row += 2

    worksheet_budget.write_row(
        row, col, ["Variable"],
        var_seperator
    )
    
    for key, value in sorted(var.items()):
        row += 1
        worksheet_budget.write_row(row, col, [key.capitalize()], normal)
        worksheet_budget.write_row(row, col + 1, [value], num_field)
            
    row += 2
    
    worksheet_budget.write_row(row, col, ["Total Variable"], normal)
    worksheet_budget.write_row(row, col + 1, [var_total], num_field)

    row += 2

    worksheet_budget.write_row(row, col, ["Total Expenses"], header)
    worksheet_budget.write_row(row, col + 1, [var_total + non_var_total], num_field)
    
    row += 2
    
    position = 'A' + str(row) + ":" + 'A' + str(row + 4)
    worksheet_budget.merge_range(position, "Monthly Income", header)

    position = 'B' + str(row) + ":" + 'B' + str(row + 4)
    worksheet_budget.merge_range(position, income, num_field)
    
    row += 5
    
    position = 'A' + str(row) + ":" + 'A' + str(row + 4)
    worksheet_budget.merge_range(position, "Extra Income after expenses", header)

    position = 'B' + str(row) + ":" + 'B' + str(row + 4)
    worksheet_budget.merge_range(position, extra_income, num_field)

    row = 0
    
    for key, value in sorted(var.items()):
        row += 1
        worksheet_budget.write_row(row, col + 2, [key.capitalize()], normal)
        worksheet_budget.write_row(row, col + 3, ['=SUMIF(F3:F66,"{}",G3:G66)'.format(key)], num_field)
        total += value
        
    row += 2

    worksheet_budget.write_row(row, col + 2, ["Total"], normal)
    worksheet_budget.write_row(row, col + 3, ["=SUM(D2:D{})".format((row-1))], num_field)

    row += 2
    temp_row = row

    position = 'C' + str(row) + ':' + 'D' + str(row)
    worksheet_budget.merge_range(position, 'Left in budget', header)

    worksheet_budget.conditional_format('D{}:D100'.format(temp_row), {'type': 'cell',
                                                                      'criteria': 'between',
                                                                      'minimum': 0,
                                                                      'maximum': 25,
                                                                      'format': caution})

    worksheet_budget.conditional_format('D{}:D100'.format(temp_row), {'type': 'cell',
                                                                      'criteria': '<',
                                                                      'value': 0,
                                                                      'format': over})

    counter = 2
    for key, value in sorted(var.items()):
        position = 'D' + str(counter)
        worksheet_budget.write_row(row, col + 2, [key.capitalize()], normal)
        worksheet_budget.write_row(row, col + 3, ['=SUM({0} - {1})'.format(value, position)], num_field)
        total += value
        row += 1
        counter += 1

    row += 2

    worksheet_budget.write_row(row, col + 2, ["Surplus"], normal)
    worksheet_budget.write_row(row, col + 3, [extra_income], num_field)

    row += 2
    worksheet_budget.write_row(row, col + 2, ["Total"], normal)
    worksheet_budget.write_row(row, col + 3, ["=SUM(D{}:D{})".format((temp_row + 1), (row - 1))], num_field)
    
    worksheet_budget.add_table('E2:H50', {'data': data,
                               'columns': [{'header': 'Item'},
                                           {'header': 'Category'},
                                           {'header': 'Cost', 'format': num_field},
                                           {'header': 'Date', 'format': date}
                                           ]})
    
    workbook.close()


if __name__ == '__main__':
    main_menu()
